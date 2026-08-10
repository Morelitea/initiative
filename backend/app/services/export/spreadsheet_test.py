"""Unit tests for the spreadsheet-document renderers.

Focus: the workbook shape. A v3 snapshot has to reach xlsx as one
worksheet per sheet, a legacy v1/v2 snapshot has to keep rendering, and
CSV — which cannot express a workbook — has to make its single-sheet
choice explicit and predictable.
"""

import io

import pytest
from openpyxl import load_workbook

from app.services.export.spreadsheet import render_csv, render_xlsx, sheets_of


def _sheet(name: str, cells: dict, **extra) -> dict:
    return {
        "id": name.lower(),
        "name": name,
        "dimensions": {"rows": 3, "cols": 2},
        "cells": cells,
        "columns": {},
        "rows": {},
        "cellStyles": {},
        "frozen": {"rows": 0, "cols": 0},
        **extra,
    }


def _workbook(*sheets: dict) -> dict:
    return {"schema_version": 3, "kind": "spreadsheet", "sheets": list(sheets)}


@pytest.mark.unit
def test_sheets_of_reads_a_v3_workbook():
    content = _workbook(_sheet("One", {}), _sheet("Two", {}))
    assert [s["name"] for s in sheets_of(content)] == ["One", "Two"]


@pytest.mark.unit
def test_sheets_of_treats_a_legacy_snapshot_as_one_sheet():
    """A document not re-saved since multi-sheet landed has its structures
    at the top level and no ``sheets`` key."""
    legacy = {"schema_version": 2, "kind": "spreadsheet", "cells": {"0:0": "x"}}
    assert sheets_of(legacy) == [legacy]


@pytest.mark.unit
def test_xlsx_renders_one_worksheet_per_sheet():
    content = _workbook(
        _sheet("Summary", {"0:0": "total", "0:1": "=SUM(Data!A1:A2)"}),
        _sheet("Data", {"0:0": 1, "1:0": 2}),
    )
    book = load_workbook(io.BytesIO(render_xlsx(content, title="Doc")))
    assert book.sheetnames == ["Summary", "Data"]
    # The cross-sheet formula travels verbatim; Excel re-evaluates it.
    assert book["Summary"]["B1"].value == "=SUM(Data!A1:A2)"
    assert book["Data"]["A1"].value == 1
    assert book["Data"]["A2"].value == 2


@pytest.mark.unit
def test_xlsx_titles_a_legacy_snapshot_from_the_document_name():
    legacy = {
        "schema_version": 2,
        "kind": "spreadsheet",
        "dimensions": {"rows": 1, "cols": 1},
        "cells": {"0:0": "x"},
    }
    book = load_workbook(io.BytesIO(render_xlsx(legacy, title="Q2: Numbers/Plan")))
    # Forbidden characters stripped, capped at Excel's 31.
    assert book.sheetnames == ["Q2 NumbersPlan"]


@pytest.mark.unit
def test_xlsx_breaks_colliding_worksheet_titles():
    """openpyxl raises on a duplicate title, and a hand-edited payload can
    still carry one even though names are de-duplicated on write."""
    content = _workbook(_sheet("Data", {}), {**_sheet("Data", {}), "id": "other"})
    book = load_workbook(io.BytesIO(render_xlsx(content, title="Doc")))
    assert book.sheetnames == ["Data", "Data 2"]


@pytest.mark.unit
def test_xlsx_keeps_per_sheet_frozen_panes():
    content = _workbook(
        _sheet("A", {}, frozen={"rows": 1, "cols": 0}),
        _sheet("B", {}, frozen={"rows": 0, "cols": 2}),
    )
    book = load_workbook(io.BytesIO(render_xlsx(content, title="Doc")))
    assert book["A"].freeze_panes == "A2"
    assert book["B"].freeze_panes == "C1"


@pytest.mark.unit
def test_csv_renders_the_first_sheet_only():
    content = _workbook(
        _sheet("First", {"0:0": "a", "0:1": "b"}),
        _sheet("Second", {"0:0": "should not appear"}),
    )
    text = render_csv(content).decode("utf-8")
    assert "a,b" in text
    assert "should not appear" not in text


@pytest.mark.unit
def test_csv_still_renders_a_legacy_snapshot():
    legacy = {
        "schema_version": 2,
        "kind": "spreadsheet",
        "dimensions": {"rows": 1, "cols": 2},
        "cells": {"0:0": "a", "0:1": "b"},
    }
    assert "a,b" in render_csv(legacy).decode("utf-8")
