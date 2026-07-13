"""Unit tests for the local Typst render backend."""

import pytest

from app.services.export.contract import RenderItem, RenderRequest
from app.services.export.local_backend import (
    LocalRenderBackend,
    UnknownTemplateError,
    resolve_template,
)

pytestmark = pytest.mark.unit


def _request(format: str = "pdf", **data) -> RenderRequest:
    payload = {
        "title": "Tasks",
        "subtitle": "unit test",
        "footer": "Tasks export",
        "columns": [
            {"key": "title", "label": "Task"},
            {"key": "status", "label": "Status"},
        ],
        "rows": [
            {
                "title": "A task",
                "project": "P",
                "status": "To Do",
                "priority": "high",
                "due": "2026-07-12",
                "assignees": "Alice",
            }
        ],
    }
    payload.update(data)
    return RenderRequest(
        guild_id=1,
        template_id="task-table",
        format=format,
        batch=(RenderItem(key="tasks", data=payload),),
    )


async def test_renders_pdf_bytes():
    artifacts = await LocalRenderBackend().render(_request())
    assert len(artifacts) == 1
    artifact = artifacts[0]
    assert artifact.key == "tasks"
    assert artifact.content_type == "application/pdf"
    assert artifact.content.startswith(b"%PDF")


async def test_user_text_is_data_not_markup():
    """Typst markup / code syntax in user strings must render as literal text,
    not execute — the sys.inputs guard. A successful compile of hostile
    strings (unbalanced markup would otherwise be a compile error) is the
    signal."""
    hostile = {
        "title": '#import "x.typ"; *bold* #sys.inputs [',
        "rows": [
            {
                "title": "#eval(1+1) ]] #footnote[x] $ broken math",
                "project": "*",
                "status": "_",
                "priority": "#",
                "due": "[",
                "assignees": "]",
            }
        ],
    }
    artifacts = await LocalRenderBackend().render(_request(**hostile))
    assert artifacts[0].content.startswith(b"%PDF")


async def test_empty_rows_renders():
    artifacts = await LocalRenderBackend().render(_request(rows=[]))
    assert artifacts[0].content.startswith(b"%PDF")


def test_template_id_is_whitelisted():
    assert resolve_template("task-table").name == "task-table.typ"
    for bad in ("../secrets", "task table", "no-such-template", "TASK-TABLE", ""):
        with pytest.raises(UnknownTemplateError):
            resolve_template(bad)


async def test_renders_csv_with_formula_neutralization():
    """CSV rides the platform exporter's safety: BOM prefix and a leading
    formula trigger prefixed so spreadsheets treat the cell as text."""
    artifacts = await LocalRenderBackend().render(
        _request(
            format="csv",
            rows=[{"title": "=HYPERLINK(evil)", "status": "To Do"}],
        )
    )
    artifact = artifacts[0]
    assert artifact.content_type.startswith("text/csv")
    text = artifact.content.decode("utf-8")
    assert text.startswith("﻿")
    assert "Task,Status" in text
    assert "'=HYPERLINK(evil)" in text


async def test_renders_xlsx_with_formula_neutralization():
    from io import BytesIO

    from openpyxl import load_workbook

    artifacts = await LocalRenderBackend().render(
        _request(
            format="xlsx",
            rows=[{"title": "=CMD|dangerous", "status": "Done"}],
        )
    )
    artifact = artifacts[0]
    assert artifact.content.startswith(b"PK")  # zip container
    sheet = load_workbook(BytesIO(artifact.content)).active
    assert [c.value for c in sheet[1]] == ["Task", "Status"]
    row = [c.value for c in sheet[2]]
    assert row == ["'=CMD|dangerous", "Done"]
    # Neutralized: stored as a string cell, not an inferred formula.
    assert sheet.cell(row=2, column=1).data_type == "s"


async def test_renders_markdown_table_with_escaped_cells():
    """Pipes and newlines in user text must not break the GFM table."""
    artifacts = await LocalRenderBackend().render(
        _request(
            format="md",
            rows=[{"title": "a | b\nmultiline", "status": "To Do"}],
        )
    )
    artifact = artifacts[0]
    assert artifact.content_type.startswith("text/markdown")
    text = artifact.content.decode("utf-8")
    lines = text.splitlines()
    assert lines[0] == "# Tasks"
    assert "| Task | Status |" in lines
    assert "| --- | --- |" in lines
    assert "| a \\| b multiline | To Do |" in lines


async def test_renders_markdown_checklist_layout():
    """layout=checklist renders GitHub-style task items — checked from the
    row's done flag, details from the non-title columns, empties skipped."""
    artifacts = await LocalRenderBackend().render(
        _request(
            format="md",
            layout="checklist",
            rows=[
                {"title": "Open item", "status": "To Do", "done": False},
                {"title": "Shipped | piped", "status": "Done", "done": True},
                {"title": "", "status": "", "done": False},
            ],
        )
    )
    text = artifacts[0].content.decode("utf-8")
    lines = text.splitlines()
    assert "- [ ] Open item (To Do)" in lines
    assert "- [x] Shipped \\| piped (Done)" in lines
    assert "- [ ] (untitled)" in lines
    assert not any("---" in line for line in lines)  # no table separator


async def test_xlsx_sanitizes_sheet_title():
    """openpyxl raises InvalidSheetTitle on []:*?/\\ — a title carrying user
    text (e.g. a guild named "My App: Dev") must render, not crash."""
    from io import BytesIO

    from openpyxl import load_workbook

    artifacts = await LocalRenderBackend().render(
        _request(format="xlsx", title="Tasks — My App: Dev [beta] a/b\\c *?")
    )
    sheet = load_workbook(BytesIO(artifacts[0].content)).active
    assert sheet.title == "Tasks — My App Dev beta abc"
    assert len(sheet.title) <= 31

    # A title reduced to nothing falls back to the item key.
    artifacts = await LocalRenderBackend().render(_request(format="xlsx", title=":::"))
    sheet = load_workbook(BytesIO(artifacts[0].content)).active
    assert sheet.title == "tasks"


async def test_xlsx_preserves_numeric_cells():
    """Neutralization must not coerce numbers to strings: ``-5`` looks like a
    formula trigger as text, but openpyxl can't infer a formula from an int,
    and a string cell would break sorting/arithmetic in the sheet."""
    from io import BytesIO

    from openpyxl import load_workbook

    artifacts = await LocalRenderBackend().render(
        _request(
            format="xlsx",
            columns=[
                {"key": "title", "label": "Task"},
                {"key": "count", "label": "Count"},
            ],
            rows=[{"title": "-starts with trigger", "count": -5}],
        )
    )
    sheet = load_workbook(BytesIO(artifacts[0].content)).active
    assert sheet.cell(row=2, column=1).value == "'-starts with trigger"
    assert sheet.cell(row=2, column=2).value == -5
    assert sheet.cell(row=2, column=2).data_type == "n"


async def test_document_template_interleaves_nested_lists():
    """A nested list must render directly beneath its parent item, not after
    all siblings — and the parent numbering must continue past the detour."""
    import io

    from pypdf import PdfReader

    request = RenderRequest(
        guild_id=1,
        template_id="document",
        format="pdf",
        batch=(
            RenderItem(
                key="doc",
                data={
                    "title": "T",
                    "blocks": [
                        {
                            "type": "list",
                            "ordered": True,
                            "checklist": False,
                            "items": [
                                {"runs": [{"text": "ALPHA"}]},
                                {
                                    "runs": [{"text": "BRAVO"}],
                                    "children": [
                                        {
                                            "type": "list",
                                            "ordered": False,
                                            "checklist": False,
                                            "items": [
                                                {"runs": [{"text": "NESTEDBRAVO"}]}
                                            ],
                                        }
                                    ],
                                },
                                {"runs": [{"text": "CHARLIE"}]},
                            ],
                        }
                    ],
                },
            ),
        ),
    )
    artifacts = await LocalRenderBackend().render(request)
    text = PdfReader(io.BytesIO(artifacts[0].content)).pages[0].extract_text()
    assert text.index("ALPHA") < text.index("BRAVO")
    assert text.index("BRAVO") < text.index("NESTEDBRAVO")
    assert text.index("NESTEDBRAVO") < text.index("CHARLIE")
    assert "3. CHARLIE" in text  # numbering survives the nested detour


async def test_document_pdf_degrades_missing_asset(monkeypatch, tmp_path):
    """Typst fails a compile on a missing image file, so an asset gone from
    storage must degrade its block to alt text — not fail the export."""
    import io

    from pypdf import PdfReader

    from app.core.config import settings

    monkeypatch.setattr(settings, "UPLOADS_DIR", str(tmp_path))
    request = RenderRequest(
        guild_id=1,
        template_id="document",
        format="pdf",
        batch=(
            RenderItem(
                key="doc",
                data={
                    "title": "T",
                    "blocks": [
                        {"type": "image", "asset": "gone.png", "alt": "GONEALT"},
                        {"type": "paragraph", "runs": [{"text": "AFTERWARDS"}]},
                    ],
                    "assets": [{"key": "gone.png", "name": "gone.png"}],
                },
            ),
        ),
    )
    artifacts = await LocalRenderBackend().render(request)
    text = PdfReader(io.BytesIO(artifacts[0].content)).pages[0].extract_text()
    assert "GONEALT" in text  # degraded to alt text
    assert "AFTERWARDS" in text  # rest of the document survived


async def test_tabular_formats_skip_template_resolution():
    """A csv/xlsx render must not require a .typ template — the payload's
    columns/rows are the whole input."""
    req = RenderRequest(
        guild_id=1,
        template_id="no-such-template",
        format="csv",
        batch=(
            RenderItem(
                key="tasks",
                data={"columns": [{"key": "title", "label": "Task"}], "rows": []},
            ),
        ),
    )
    artifacts = await LocalRenderBackend().render(req)
    assert artifacts[0].content_type.startswith("text/csv")


async def test_renders_markdown_numbered_layout():
    """layout=numbered renders an ordered turn-order list — details from the
    non-title columns, the ``current`` row bolded, ``order`` never doubled
    into the detail trail."""
    artifacts = await LocalRenderBackend().render(
        _request(
            format="md",
            layout="numbered",
            columns=[
                {"key": "order", "label": "#"},
                {"key": "title", "label": "Item"},
                {"key": "member", "label": "Member"},
                {"key": "status", "label": "Status"},
            ],
            rows=[
                # member=None must be skipped BEFORE stringifying — str(None)
                # passes a truthiness guard, and its empty rendering would
                # smuggle a spurious "( · Current)" into the detail trail.
                {
                    "order": 1,
                    "title": "Alice | piped",
                    "member": None,
                    "status": "Current",
                    "current": True,
                },
                {"order": 2, "title": "Bob", "member": "", "status": ""},
                {"order": 3, "title": "", "member": None, "status": "Hidden"},
            ],
        )
    )
    text = artifacts[0].content.decode("utf-8")
    lines = text.splitlines()
    assert "1. **Alice \\| piped** (Current)" in lines
    assert "2. Bob" in lines
    assert "3. (untitled) (Hidden)" in lines
    assert not any("---" in line for line in lines)  # no table separator


async def test_data_table_template_renders_payload_columns():
    """The generic template's whole point: the column set (labels, order,
    numeric cells) comes from the payload, no bespoke .typ per source."""
    import io

    from pypdf import PdfReader

    request = RenderRequest(
        guild_id=1,
        template_id="data-table",
        format="pdf",
        batch=(
            RenderItem(
                key="counters",
                data={
                    "title": "Party Resources",
                    "subtitle": "unit test",
                    "footer": "counters export",
                    "description": "Session 12 snapshot",
                    "columns": [
                        {"key": "title", "label": "Counter", "width": "2fr"},
                        {"key": "count", "label": "Count"},
                        {"key": "max", "label": "Max"},
                    ],
                    "rows": [
                        {"title": "Torches", "count": 5, "max": 10},
                        {"title": "Rations", "count": 2.5, "max": ""},
                    ],
                },
            ),
        ),
    )
    artifacts = await LocalRenderBackend().render(request)
    assert artifacts[0].content.startswith(b"%PDF")
    # Compare space-insensitively: pypdf's naive extraction inserts a space at
    # Outfit-kerned pairs (``Torches`` -> ``T orches``); the render is correct.
    text = PdfReader(io.BytesIO(artifacts[0].content)).pages[0].extract_text()
    packed = text.replace(" ", "")
    assert "PartyResources" in packed
    assert "Session12snapshot" in packed
    assert "Counter" in packed and "Torches" in packed
    assert "2.5" in packed  # numeric cells render as text


async def test_data_table_template_survives_hostile_text_and_empty_rows():
    """sys.inputs guard holds for the generic template too: Typst markup in
    user strings stays data, and zero rows still compile."""
    hostile = RenderRequest(
        guild_id=1,
        template_id="data-table",
        format="pdf",
        batch=(
            RenderItem(
                key="q",
                data={
                    "title": '#import "x.typ"; *bold* [',
                    "description": "#eval(1+1) ]] $ broken math",
                    "columns": [{"key": "title", "label": "*"}],
                    "rows": [{"title": "#sys.inputs ["}],
                },
            ),
        ),
    )
    artifacts = await LocalRenderBackend().render(hostile)
    assert artifacts[0].content.startswith(b"%PDF")

    empty = RenderRequest(
        guild_id=1,
        template_id="data-table",
        format="pdf",
        batch=(RenderItem(key="q", data={"title": "T", "columns": [], "rows": []}),),
    )
    artifacts = await LocalRenderBackend().render(empty)
    assert artifacts[0].content.startswith(b"%PDF")


async def test_report_pdf_embeds_outfit_font():
    """Reports render in Outfit (the web UI's typeface), bundled and staged via
    font_paths — so the PDF matches the app and doesn't depend on host fonts.
    A subset per weight is embedded as ``<TAG>+Outfit-<Weight>``."""
    import re

    artifacts = await LocalRenderBackend().render(_request())
    base_fonts = re.findall(rb"/BaseFont\s*/([A-Za-z0-9+\-]+)", artifacts[0].content)
    families = {name.decode().split("+", 1)[-1] for name in base_fonts}
    # Regular body + bold headers/labels both come from the bundled Outfit cuts.
    assert "Outfit-Regular" in families
    assert "Outfit-Bold" in families
