"""Spreadsheet-document renderers: the sparse cell map as CSV or styled XLSX.

Input is the canonical v2 snapshot ``documents_spreadsheet`` persists:
``{dimensions, cells, columns, rows, cellStyles, frozen}`` with ``"r:c"``
keys. CSV carries values (BOM-prefixed); XLSX additionally maps the
formatting model — fonts, colors, alignment, borders, number formats, column
widths, row heights, frozen panes.

Formulas: the app's sheets store formulas as ``=``-prefixed cell strings
(evaluated client-side; the snapshot keeps the raw text). Unlike the tabular
exports — where a leading ``=`` is FOREIGN text smuggling a formula — a
spreadsheet document's ``=`` cells are the user's own first-class content,
so grid exports preserve them (Excel re-evaluates; the app's function set is
an Excel subset). Mirrors the frontend's ``isFormula`` exactly: ``=`` prefix
only. Other trigger prefixes (``+ - @``) are still neutralized in CSV (they
are smuggling vectors, never app formulas); XLSX needs no guard for them
(openpyxl only ever infers a formula from a ``=`` string).

Style precedence mirrors the editor: column style, then row style, then the
cell's own style, later layers winning per key.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.platform.csv_export import neutralize_cell

# Frontend sizes are CSS pixels; Excel wants points (rows) and its own
# character-width unit (columns, ~7px per unit at the default font).
_PX_TO_POINTS = 0.75
_PX_PER_WIDTH_UNIT = 7.0

_VALIGN_MAP = {"top": "top", "middle": "center", "bottom": "bottom"}

_DATE_FORMATS = {"iso": "yyyy-mm-dd", "us": "mm/dd/yyyy", "eu": "dd/mm/yyyy"}


def _cell_coords(key: str) -> tuple[int, int] | None:
    """Parse an ``"r:c"`` cell key; None for anything malformed. The
    normalizer enforces the shape on write, but a renderer must not 500 on a
    corrupted snapshot — bad keys are simply skipped."""
    parts = key.split(":") if isinstance(key, str) else []
    if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
        return None
    return int(parts[0]), int(parts[1])


def _grid_size(content: dict) -> tuple[int, int]:
    dimensions = content.get("dimensions") or {}
    rows = int(dimensions.get("rows") or 0)
    cols = int(dimensions.get("cols") or 0)
    for key in content.get("cells") or {}:
        coords = _cell_coords(key)
        if coords is None:
            continue
        rows = max(rows, coords[0] + 1)
        cols = max(cols, coords[1] + 1)
    return rows, cols


def _csv_cell(value: object) -> object:
    """The user's own ``=`` formulas pass through (first-class content, and
    Excel evaluates them from CSV too); everything else keeps the platform
    exporter's trigger neutralization."""
    if isinstance(value, str) and value.startswith("="):
        return value
    return neutralize_cell(value)


def render_csv(content: dict) -> bytes:
    rows, cols = _grid_size(content)
    cells = content.get("cells") or {}
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for r in range(rows):
        writer.writerow([_csv_cell(cells.get(f"{r}:{c}")) for c in range(cols)])
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def render_xlsx(content: dict, *, title: str) -> bytes:
    rows, cols = _grid_size(content)
    cells = content.get("cells") or {}
    columns = content.get("columns") or {}
    row_fmts = content.get("rows") or {}
    cell_styles = content.get("cellStyles") or {}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title(title)

    for r in range(rows):
        for c in range(cols):
            # Raw values: a ``=`` string becomes a live formula (openpyxl
            # data_type 'f'), everything else stays a typed value — see the
            # module docstring for why the grid does NOT neutralize.
            value = cells.get(f"{r}:{c}")
            cell = sheet.cell(row=r + 1, column=c + 1)
            if value is not None:
                cell.value = value
            style = _effective_style(columns, row_fmts, cell_styles, r, c)
            if style:
                _apply_style(cell, style)

    for key, entry in columns.items():
        width = entry.get("width") if isinstance(entry, dict) else None
        if isinstance(width, (int, float)):
            letter = get_column_letter(int(key) + 1)
            sheet.column_dimensions[letter].width = width / _PX_PER_WIDTH_UNIT
    for key, entry in row_fmts.items():
        height = entry.get("height") if isinstance(entry, dict) else None
        if isinstance(height, (int, float)):
            sheet.row_dimensions[int(key) + 1].height = height * _PX_TO_POINTS

    frozen = content.get("frozen") or {}
    frozen_rows = int(frozen.get("rows") or 0)
    frozen_cols = int(frozen.get("cols") or 0)
    if frozen_rows or frozen_cols:
        sheet.freeze_panes = sheet.cell(row=frozen_rows + 1, column=frozen_cols + 1)

    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def _sheet_title(title: str) -> str:
    cleaned = title.translate(str.maketrans("", "", "[]:*?/\\")).strip()
    return cleaned[:31] or "Sheet1"


def _effective_style(
    columns: dict, row_fmts: dict, cell_styles: dict, r: int, c: int
) -> dict[str, Any]:
    """Merge column -> row -> cell styles, later layers winning per key."""
    merged: dict[str, Any] = {}
    for layer in (
        (columns.get(str(c)) or {}).get("style"),
        (row_fmts.get(str(r)) or {}).get("style"),
        cell_styles.get(f"{r}:{c}"),
    ):
        if isinstance(layer, dict):
            merged.update(layer)
    return merged


def _apply_style(cell, style: dict[str, Any]) -> None:
    font_kwargs: dict[str, Any] = {}
    if style.get("bold"):
        font_kwargs["bold"] = True
    if style.get("italic"):
        font_kwargs["italic"] = True
    if style.get("underline"):
        font_kwargs["underline"] = "single"
    if style.get("strike"):
        font_kwargs["strike"] = True
    color = _argb(style.get("color"))
    if color:
        font_kwargs["color"] = color
    font_size = style.get("fontSize")
    if isinstance(font_size, (int, float)):
        font_kwargs["size"] = round(font_size * _PX_TO_POINTS, 1)
    if font_kwargs:
        cell.font = Font(**font_kwargs)

    fill = _argb(style.get("fill"))
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")

    align = style.get("align")
    valign = _VALIGN_MAP.get(style.get("valign") or "")
    if align or valign:
        cell.alignment = Alignment(horizontal=align, vertical=valign)

    border = style.get("border")
    if isinstance(border, dict):
        sides: dict[str, Side] = {}
        for edge in ("top", "right", "bottom", "left"):
            spec = border.get(edge)
            if isinstance(spec, dict):
                sides[edge] = Side(
                    style=spec.get("style", "thin"),
                    color=_argb(spec.get("color")) or "FF000000",
                )
        if sides:
            cell.border = Border(**sides)

    fmt = style.get("format")
    if isinstance(fmt, dict):
        number_format = _number_format(fmt)
        if number_format:
            cell.number_format = number_format


def _argb(value: Any) -> str | None:
    """``#rrggbb`` (or ``#rgb`` shorthand) -> openpyxl's 8-char ARGB, else
    None. openpyxl raises on any other length, and a renderer must not 500
    over one bad formatting entry — invalid colors are simply dropped."""
    if not isinstance(value, str) or not value.startswith("#"):
        return None
    hex_part = value[1:]
    if len(hex_part) == 3:
        hex_part = "".join(ch * 2 for ch in hex_part)
    if len(hex_part) != 6 or any(c not in "0123456789abcdefABCDEF" for c in hex_part):
        return None
    return f"FF{hex_part.upper()}"


def _number_format(fmt: dict[str, Any]) -> str | None:
    ftype = fmt.get("type")
    decimals = int(fmt.get("decimals") or 0)
    places = f".{'0' * decimals}" if decimals else ""
    grouping = "#,##0" if fmt.get("grouping", True) else "0"
    if ftype == "percent":
        return f"0{places}%"
    if ftype == "date":
        return _DATE_FORMATS.get(fmt.get("pattern") or "iso", "yyyy-mm-dd")
    if ftype == "fixed":
        return f"{grouping}{places}"
    if ftype == "currency":
        code = str(fmt.get("currency") or "USD")
        base = f'{grouping}{places}\\ "{code}"'
        if fmt.get("negatives") in ("red", "redParens"):
            return f"{base};[Red]-{base}"
        return base
    return None  # "plain" and unknowns: leave the default General format
